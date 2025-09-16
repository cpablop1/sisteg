from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import TipoPago
from apps.compra.models import Compra, DetalleCompra
from apps.servicio.models import Servicio, Garantia, TipoServicio
from apps.producto.models import Producto

@login_required(login_url='autenticacion')
def vista_inicio(request):
    return render(request, 'inicio/inicio.html')

@login_required(login_url='autenticacion')
def listar_tipo_pago(request):
    # Mensajes de respuesta
    res = False
    msg = 'Error al listar tipos de pagos.'
    data = {}
    data['data'] = []
    data["page_range"] = []
    id = request.GET.get('id', None) or None
    buscar = request.GET.get('buscar', '').strip() or ''
    pagina = request.GET.get('pagina', 1) or 1
    select = request.GET.get('select', None) or None
    tipo_pago = ''

    try:
        if id: # Verificamos si necesitamos una categoría expecífica
            tipo_pago = TipoPago.objects.filter(id = id)
        elif len(buscar) > 0: # Verificamos si hay búsquedad
            tipo_pago = TipoPago.objects.filter(
                Q(descripcion__icontains = buscar) # Si hay buscamos por descripción
            )
        else:
            # Obtenemos todas las categorías
            tipo_pago = TipoPago.objects.all()
            
        if select:
            paginas = tipo_pago
        else:
            # Paginamos las categorías
            paginador = Paginator(tipo_pago, 10)
            # Obtenemos la página
            paginas = paginador.get_page(pagina)
            # Preparamos el listado
        for tip in paginas:
            data['data'].append(
                {
                    'id': tip.id,
                    'descripcion': tip.descripcion,
                    'fecha_ingreso': tip.fecha_ingreso,
                    'fecha_actualizacion': tip.fecha_actualizacion,
                    'usuario': tip.usuario_id.username
                }
            )
        # Preparamos la visualización de las páginas
        if paginador.num_pages > 5:
            start = int(pagina)
            end = int(pagina) + 5
            if end > paginador.num_pages:
                start = paginador.num_pages - 4
                end = paginador.num_pages + 1
            for i in range(start, end):
                data["page_range"].append(i)
        else:
            for i in range(paginador.num_pages):
                data["page_range"].append(i + 1)
        data["num_pages"] = paginador.num_pages
        data["has_next"] = paginas.has_next()
        data["has_previous"] = paginas.has_previous()
        data["count"] = paginador.count

        # Preparamos mensajes de respuesta
        res = True
        msg = 'Listado de listado de tipo de pagos.'
    except:
        res = False

    data['res'] = res
    data['msg'] = msg
    # Retornamos los datos
    return JsonResponse(data)

def estadistica(request):
    compra = ''
    servicio = ''
    producto  = ''
    data = {}
    data['compra'] = []
    data['servicio'] = []
    data['grafica'] = []

    try:
        # Preparamos los datos de compra
        compra = Compra.objects.all()
        total = sum(co.subtotal for co in compra)
        cantidad = compra.count()
        data['compra'].append({
            'total': f'Q {total} en compras',
            'cantidad': f'{cantidad} realizadas'
        })
        # Preparamos los datos de servicio
        tipo_servicio = TipoServicio.objects.all()
        for tp in tipo_servicio:
            servicio = Servicio.objects.filter(tipo_servicio_id = tp.id)
            total = sum(ser.subtotal for ser in servicio)
            cantidad = servicio.count()
            data['servicio'].append({
                'tipo_servicio': tp.descripcion,
                'total': f'Q {total} en {tp.descripcion}',
                'cantidad': f'{cantidad} realizadas'
            })
        # Preparamos los datos para la gráfica
        for tp in tipo_servicio:
            servicio = Servicio.objects.filter(tipo_servicio_id = tp.id)
            cantidad = servicio.count()
            data['grafica'].append({
                'descripcion': tp.descripcion,
                'cantidad': cantidad
            })
    except:
        pass

    return JsonResponse(data)

@login_required(login_url='autenticacion')
def reporte_completo_excel(request):
    """
    Genera un reporte completo del sistema en formato Excel.
    Incluye: Compras, Servicios, Productos y Estadísticas generales.
    
    Returns:
        HttpResponse: Archivo Excel con el reporte completo
    """
    try:
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, color="000000", size=14)
        data_font = Font(size=10)
        
        # Colores para diferentes secciones
        colors = {
            'compras': '366092',
            'servicios': 'ffc107', 
            'productos': 'dc3545',
            'estadisticas': '28a745'
        }
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        # Borde
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. HOJA DE COMPRAS
        ws_compras = wb.active
        ws_compras.title = "Compras"
        
        # Título
        ws_compras.merge_cells('A1:H1')
        ws_compras['A1'] = "REPORTE DE COMPRAS"
        ws_compras['A1'].font = title_font
        ws_compras['A1'].alignment = center_alignment
        ws_compras['A1'].fill = PatternFill(start_color=colors['compras'], end_color=colors['compras'], fill_type="solid")
        
        # Encabezados de compras
        headers_compras = ["ID", "Proveedor", "Tipo de Pago", "Subtotal", "Fecha de Ingreso", "Usuario", "Estado"]
        for col, header in enumerate(headers_compras, 1):
            cell = ws_compras.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color=colors['compras'], end_color=colors['compras'], fill_type="solid")
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Datos de compras
        compras = Compra.objects.select_related('proveedor_id', 'tipo_pago_id', 'usuario_id').all()
        for row, compra in enumerate(compras, 4):
            ws_compras.cell(row=row, column=1, value=compra.id)
            ws_compras.cell(row=row, column=2, value=f"{compra.proveedor_id.nombres} {compra.proveedor_id.apellidos}")
            ws_compras.cell(row=row, column=3, value=compra.tipo_pago_id.descripcion)
            ws_compras.cell(row=row, column=4, value=f"Q {compra.subtotal:.2f}")
            ws_compras.cell(row=row, column=5, value=compra.fecha_ingreso.strftime("%d/%m/%Y %H:%M"))
            ws_compras.cell(row=row, column=6, value=compra.usuario_id.username)
            ws_compras.cell(row=row, column=7, value="Confirmada" if compra.estado else "Pendiente")
            
            # Aplicar bordes a todas las celdas
            for col in range(1, len(headers_compras) + 1):
                ws_compras.cell(row=row, column=col).border = thin_border
                ws_compras.cell(row=row, column=col).font = data_font
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers_compras) + 1):
            column_letter = get_column_letter(col)
            ws_compras.column_dimensions[column_letter].width = 20
        
        # 2. HOJA DE SERVICIOS
        ws_servicios = wb.create_sheet("Servicios")
        
        # Título
        ws_servicios.merge_cells('A1:G1')
        ws_servicios['A1'] = "REPORTE DE SERVICIOS FINALIZADOS"
        ws_servicios['A1'].font = title_font
        ws_servicios['A1'].alignment = center_alignment
        ws_servicios['A1'].fill = PatternFill(start_color=colors['servicios'], end_color=colors['servicios'], fill_type="solid")
        
        # Encabezados de servicios
        headers_servicios = ["ID", "Cliente", "Usuario", "Tipo de Servicio", "Subtotal", "Fecha de Ingreso", "Estado"]
        for col, header in enumerate(headers_servicios, 1):
            cell = ws_servicios.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color=colors['servicios'], end_color=colors['servicios'], fill_type="solid")
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Datos de servicios (solo finalizados)
        servicios = Servicio.objects.select_related(
            'cliente_id', 'tipo_servicio_id', 'usuario_id'
        ).filter(estado=True).all()
        
        for row, servicio in enumerate(servicios, 4):
            ws_servicios.cell(row=row, column=1, value=servicio.id)
            ws_servicios.cell(row=row, column=2, value=f"{servicio.cliente_id.nombres} {servicio.cliente_id.apellidos}")
            ws_servicios.cell(row=row, column=3, value=servicio.usuario_id.username)
            ws_servicios.cell(row=row, column=4, value=servicio.tipo_servicio_id.descripcion)
            ws_servicios.cell(row=row, column=5, value=f"Q {servicio.subtotal:.2f}")
            ws_servicios.cell(row=row, column=6, value=servicio.fecha_ingreso.strftime("%d/%m/%Y %H:%M"))
            ws_servicios.cell(row=row, column=7, value="Confirmado" if servicio.estado else "Pendiente")
            
            # Aplicar bordes a todas las celdas
            for col in range(1, len(headers_servicios) + 1):
                ws_servicios.cell(row=row, column=col).border = thin_border
                ws_servicios.cell(row=row, column=col).font = data_font
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers_servicios) + 1):
            column_letter = get_column_letter(col)
            ws_servicios.column_dimensions[column_letter].width = 20
        
        # 3. HOJA DE PRODUCTOS
        ws_productos = wb.create_sheet("Productos")
        
        # Título
        ws_productos.merge_cells('A1:L1')
        ws_productos['A1'] = "REPORTE DE PRODUCTOS (STOCK > 0)"
        ws_productos['A1'].font = title_font
        ws_productos['A1'].alignment = center_alignment
        ws_productos['A1'].fill = PatternFill(start_color=colors['productos'], end_color=colors['productos'], fill_type="solid")
        
        # Encabezados de productos
        headers_productos = ["ID", "Descripción", "Marca", "Categoría", "Costo", "Precio", "Stock", "Inversión", "Ganancia Unitaria", "Total Ganancia", "Fecha de Ingreso", "Usuario"]
        for col, header in enumerate(headers_productos, 1):
            cell = ws_productos.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color=colors['productos'], end_color=colors['productos'], fill_type="solid")
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Datos de productos (solo con stock > 0)
        productos = Producto.objects.select_related(
            'marca_id', 'categoria_id', 'usuario_id'
        ).filter(stock__gt=0).all()
        
        for row, producto in enumerate(productos, 4):
            # Calcular valores
            inversion = producto.costo * producto.stock
            ganancia_unitaria = producto.precio - producto.costo
            total_ganancia = ganancia_unitaria * producto.stock
            
            ws_productos.cell(row=row, column=1, value=producto.id)
            ws_productos.cell(row=row, column=2, value=producto.descripcion)
            ws_productos.cell(row=row, column=3, value=producto.marca_id.descripcion)
            ws_productos.cell(row=row, column=4, value=producto.categoria_id.descripcion)
            ws_productos.cell(row=row, column=5, value=f"Q {producto.costo:.2f}")
            ws_productos.cell(row=row, column=6, value=f"Q {producto.precio:.2f}")
            ws_productos.cell(row=row, column=7, value=producto.stock)
            ws_productos.cell(row=row, column=8, value=f"Q {inversion:.2f}")
            ws_productos.cell(row=row, column=9, value=f"Q {ganancia_unitaria:.2f}")
            ws_productos.cell(row=row, column=10, value=f"Q {total_ganancia:.2f}")
            ws_productos.cell(row=row, column=11, value=producto.fecha_ingreso.strftime("%d/%m/%Y %H:%M"))
            ws_productos.cell(row=row, column=12, value=producto.usuario_id.username)
            
            # Aplicar bordes a todas las celdas
            for col in range(1, len(headers_productos) + 1):
                ws_productos.cell(row=row, column=col).border = thin_border
                ws_productos.cell(row=row, column=col).font = data_font
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers_productos) + 1):
            column_letter = get_column_letter(col)
            ws_productos.column_dimensions[column_letter].width = 20
        
        # 4. HOJA DE ESTADÍSTICAS
        ws_estadisticas = wb.create_sheet("Estadísticas")
        
        # Título
        ws_estadisticas.merge_cells('A1:F1')
        ws_estadisticas['A1'] = "ESTADÍSTICAS GENERALES DEL SISTEMA POR MES (SERVICIOS FINALIZADOS)"
        ws_estadisticas['A1'].font = title_font
        ws_estadisticas['A1'].alignment = center_alignment
        ws_estadisticas['A1'].fill = PatternFill(start_color=colors['estadisticas'], end_color=colors['estadisticas'], fill_type="solid")
        
        # Encabezados de estadísticas
        headers_estadisticas = ["Mes", "Compras", "Servicios", "Ganancias", "Cantidad Compras", "Cantidad Servicios"]
        for col, header in enumerate(headers_estadisticas, 1):
            cell = ws_estadisticas.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color=colors['estadisticas'], end_color=colors['estadisticas'], fill_type="solid")
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Obtener datos agrupados por mes
        from django.db.models import Sum, Count
        from django.db.models.functions import TruncMonth
        
        # Estadísticas de compras por mes
        compras_por_mes = compras.annotate(
            mes=TruncMonth('fecha_ingreso')
        ).values('mes').annotate(
            total_compras=Sum('subtotal'),
            cantidad_compras=Count('id')
        ).order_by('mes')
        
        # Estadísticas de servicios por mes (usando ganancia real de DetalleServicio)
        from apps.servicio.models import DetalleServicio
        
        servicios_por_mes = servicios.annotate(
            mes=TruncMonth('fecha_ingreso')
        ).values('mes').annotate(
            total_servicios=Sum('subtotal'),
            total_costo_servicio=Sum('costo_servicio'),
            cantidad_servicios=Count('id')
        ).order_by('mes')
        
        # Obtener ganancias reales por mes desde DetalleServicio
        ganancias_por_mes = DetalleServicio.objects.filter(
            servicio_id__in=servicios.values_list('id', flat=True)
        ).annotate(
            mes=TruncMonth('servicio_id__fecha_ingreso')
        ).values('mes').annotate(
            total_ganancia=Sum('ganancia')
        ).order_by('mes')
        
        # Crear diccionario para combinar datos
        datos_por_mes = {}
        
        # Procesar compras
        for compra_mes in compras_por_mes:
            mes = compra_mes['mes'].strftime('%Y-%m')
            datos_por_mes[mes] = {
                'mes_nombre': compra_mes['mes'].strftime('%B %Y'),
                'compras': compra_mes['total_compras'] or 0,
                'cantidad_compras': compra_mes['cantidad_compras'],
                'servicios': 0,
                'costo_servicio': 0,
                'cantidad_servicios': 0,
                'ganancia_real': 0
            }
        
        # Procesar servicios
        for servicio_mes in servicios_por_mes:
            mes = servicio_mes['mes'].strftime('%Y-%m')
            if mes in datos_por_mes:
                datos_por_mes[mes]['servicios'] = servicio_mes['total_servicios'] or 0
                datos_por_mes[mes]['costo_servicio'] = servicio_mes['total_costo_servicio'] or 0
                datos_por_mes[mes]['cantidad_servicios'] = servicio_mes['cantidad_servicios']
            else:
                datos_por_mes[mes] = {
                    'mes_nombre': servicio_mes['mes'].strftime('%B %Y'),
                    'compras': 0,
                    'cantidad_compras': 0,
                    'servicios': servicio_mes['total_servicios'] or 0,
                    'costo_servicio': servicio_mes['total_costo_servicio'] or 0,
                    'cantidad_servicios': servicio_mes['cantidad_servicios']
                }
        
        # Procesar ganancias reales
        for ganancia_mes in ganancias_por_mes:
            mes = ganancia_mes['mes'].strftime('%Y-%m')
            if mes in datos_por_mes:
                datos_por_mes[mes]['ganancia_real'] = ganancia_mes['total_ganancia'] or 0
            else:
                datos_por_mes[mes] = {
                    'mes_nombre': ganancia_mes['mes'].strftime('%B %Y'),
                    'compras': 0,
                    'cantidad_compras': 0,
                    'servicios': 0,
                    'costo_servicio': 0,
                    'cantidad_servicios': 0,
                    'ganancia_real': ganancia_mes['total_ganancia'] or 0
                }
        
        # Escribir datos por mes
        row = 4
        for mes_key in sorted(datos_por_mes.keys()):
            datos = datos_por_mes[mes_key]
            # Usar ganancia real de DetalleServicio + costo_servicio
            ganancia_real = datos.get('ganancia_real', 0)
            costo_servicio = datos.get('costo_servicio', 0)
            ganancia_total = ganancia_real + costo_servicio
            
            ws_estadisticas.cell(row=row, column=1, value=datos['mes_nombre'])
            ws_estadisticas.cell(row=row, column=2, value=f"Q {datos['compras']:.2f}")
            ws_estadisticas.cell(row=row, column=3, value=f"Q {datos['servicios']:.2f}")
            ws_estadisticas.cell(row=row, column=4, value=f"Q {ganancia_total:.2f}")
            ws_estadisticas.cell(row=row, column=5, value=datos['cantidad_compras'])
            ws_estadisticas.cell(row=row, column=6, value=datos['cantidad_servicios'])
            
            # Aplicar bordes a todas las celdas
            for col in range(1, len(headers_estadisticas) + 1):
                ws_estadisticas.cell(row=row, column=col).border = thin_border
                ws_estadisticas.cell(row=row, column=col).font = data_font
            
            row += 1
        
        # Agregar fila de totales
        total_compras = sum(datos['compras'] for datos in datos_por_mes.values())
        total_servicios = sum(datos['servicios'] for datos in datos_por_mes.values())
        # Calcular total de ganancias reales
        total_ganancias = sum(
            (datos.get('ganancia_real', 0) + datos.get('costo_servicio', 0)) 
            for datos in datos_por_mes.values()
        )
        total_cantidad_compras = sum(datos['cantidad_compras'] for datos in datos_por_mes.values())
        total_cantidad_servicios = sum(datos['cantidad_servicios'] for datos in datos_por_mes.values())
        
        # Fila de totales con estilo diferente
        ws_estadisticas.cell(row=row, column=1, value="TOTAL GENERAL")
        ws_estadisticas.cell(row=row, column=2, value=f"Q {total_compras:.2f}")
        ws_estadisticas.cell(row=row, column=3, value=f"Q {total_servicios:.2f}")
        ws_estadisticas.cell(row=row, column=4, value=f"Q {total_ganancias:.2f}")
        ws_estadisticas.cell(row=row, column=5, value=total_cantidad_compras)
        ws_estadisticas.cell(row=row, column=6, value=total_cantidad_servicios)
        
        # Aplicar estilo especial a la fila de totales
        for col in range(1, len(headers_estadisticas) + 1):
            cell = ws_estadisticas.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers_estadisticas) + 1):
            column_letter = get_column_letter(col)
            ws_estadisticas.column_dimensions[column_letter].width = 20
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_completo_sistema_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error al generar reporte completo: {str(e)}", status=500)